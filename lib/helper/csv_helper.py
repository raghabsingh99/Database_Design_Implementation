import openpyxl
from openpyxl.utils import range_boundaries
import re


def get_merged_cell_value(sheet, row, col):
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return sheet.cell(min_row, min_col).value
    return sheet.cell(row, col).value

def to_snake_case(s):
    if s is not None:
        s = re.sub('([A-Z]+)', r'_\1', s).lower()
        s = re.sub('[^0-9a-zA-Z]+', '_', s).strip('_')
        return s
    return None

def read_excel_file(file_path: str, solution: str = '') -> list[dict]:
    workbook = openpyxl.load_workbook(file_path)
    
    
    all_data = []

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
               
        header_row = get_header_starting_line(solution=solution)
        
        
        # Read headers
        headers = []
        max_column = get_max_sheet_column(solution=solution, sheet=sheet)
        for col in range(1, max_column + 1):
            
            header = None
            if solution != "RAD":
                header = get_merged_cell_value(sheet, header_row, col)
            else:
                header = sheet.cell(row=header_row, column=col).value
                
            headers.append(to_snake_case(header))
            headers = list(filter(None, headers))
        
            
        # Read data rows into a list of dictionaries
        for row in range(header_row + 1, sheet.max_row + 1):
            
            row_data = {}
            for col in range(1, headers.__len__() + 1):
                
                header = headers[col - 1]
                value = get_merged_cell_value(sheet, row, col)
                
                row_data[header] = value
            all_data.append(row_data)
            
    workbook.close()
    return all_data

 
def get_max_sheet_column(solution: str, sheet):
    if solution == "RAD":
        return 12
    else:
        return sheet.max_column
                       
                    
def get_header_starting_line(solution) -> int:
    if solution == 'MSK' or solution == 'SLEEP' or solution == 'REHAB':
        return 4
    elif solution == "RADIOLOGY":
        return 3
    elif solution == "RAD":
        return 2
    else:
        return 1